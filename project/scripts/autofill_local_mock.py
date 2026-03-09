from pathlib import Path
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time

BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / 'data' / 'sample_input.xlsx'
LOGIN_URL = 'http://127.0.0.1:8000/mock_site/sign_in.html'
PAYMENT_URL = 'http://127.0.0.1:8000/mock_site/face_to_face_payment.html'


def read_first_row(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path)
    ws = wb['input']
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    return dict(zip(headers, values))


def create_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def login_mock_site(driver: webdriver.Chrome, row: dict) -> None:
    driver.get(LOGIN_URL)
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.ID, 'username'))).send_keys(str(row['username']))
    driver.find_element(By.ID, 'password').send_keys(str(row['password']))
    driver.find_element(By.ID, 'pin').send_keys(str(row['pin']))
    driver.find_element(By.ID, 'loginBtn').click()
    wait.until(EC.url_contains('face_to_face_payment.html'))


def fill_payment_form(driver: webdriver.Chrome, row: dict) -> None:
    wait = WebDriverWait(driver, 10)
    if 'face_to_face_payment.html' not in driver.current_url:
        driver.get(PAYMENT_URL)

    wait.until(EC.presence_of_element_located((By.ID, 'cardType')))
    Select(driver.find_element(By.ID, 'cardType')).select_by_visible_text(str(row['card_type']))

    card_number = str(row['card_number']).replace('-', '').strip()
    driver.find_element(By.ID, 'cardNumber').send_keys(card_number)
    Select(driver.find_element(By.ID, 'expMonth')).select_by_visible_text(str(row['exp_month']).zfill(2))
    Select(driver.find_element(By.ID, 'expYear')).select_by_visible_text(str(row['exp_year']).zfill(2))
    driver.find_element(By.ID, 'cardPw2').send_keys(str(row['card_password_two_digits']).zfill(2))
    Select(driver.find_element(By.ID, 'installment')).select_by_visible_text(str(row['installment']))
    driver.find_element(By.ID, 'phone').send_keys(str(row['phone']))
    driver.find_element(By.ID, 'customerName').send_keys(str(row['customer_name']))
    driver.find_element(By.ID, 'birth6').send_keys(str(row['birth6']))

    # 안전상 실제 제출은 하지 않고, 데모 버튼만 클릭합니다.
    driver.find_element(By.ID, 'previewBtn').click()
    wait.until(EC.visibility_of_element_located((By.ID, 'resultBox')))


def main() -> None:
    row = read_first_row(EXCEL_PATH)
    driver = create_driver()
    try:
        login_mock_site(driver, row)
        fill_payment_form(driver, row)
        print('Done. Review the browser window for the filled form and preview result.')
        time.sleep(5)
    finally:
        driver.quit()


if __name__ == '__main__':
    main()
